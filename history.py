# -*- coding: utf-8 -*-
"""
Histórico de eventos concluídos, salvo em uma planilha Excel
(data/historico_eventos.xlsx). Cada linha é um evento concluído, com uma
coluna por função mostrando quem participou.
"""

import datetime as dt
import os

from openpyxl import Workbook, load_workbook

import config

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
HISTORY_FILE = os.path.join(DATA_DIR, "historico_eventos.xlsx")

BASE_HEADERS = [
    "Data de conclusão",
    "Título",
    "Descrição",
    "Tamanho",
    "Início",
    "Organizador",
    "Concluído por",
    "Total confirmados",
]

# As funções são as mesmas em todos os tamanhos de evento (só a capacidade
# muda) — por isso é seguro usar o template "25" como referência das colunas.
ROLE_LABELS = [r["label"] for r in config.ROLE_TEMPLATES["25"]]

FULL_HEADERS = BASE_HEADERS + ROLE_LABELS + ["Não vou"]


def _ensure_workbook():
    os.makedirs(DATA_DIR, exist_ok=True)
    if os.path.exists(HISTORY_FILE):
        wb = load_workbook(HISTORY_FILE)
        ws = wb.active
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico"
    ws.append(FULL_HEADERS)
    for col_idx, header in enumerate(FULL_HEADERS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
            14, min(40, len(header) + 4)
        )
    return wb, ws


def add_event_record(event: dict, role_names: dict, nao_vou_names: list, concluded_at_str: str, concluded_by: str):
    """
    Adiciona uma linha ao histórico.

    role_names: dict {role_key: [nomes dos participantes ativos]}
    nao_vou_names: lista de nomes de quem marcou "não vou"
    """
    wb, ws = _ensure_workbook()

    total_confirmados = sum(len(v) for v in role_names.values())

    row = [
        concluded_at_str,
        event.get("title", ""),
        event.get("description", ""),
        event.get("max_participants", ""),
        event.get("start", ""),
        event.get("creator_name", ""),
        concluded_by,
        total_confirmados,
    ]

    role_label_by_key = {r["key"]: r["label"] for r in event["roles_config"]}
    for label in ROLE_LABELS:
        key = next((k for k, lbl in role_label_by_key.items() if lbl == label), None)
        names = role_names.get(key, []) if key else []
        row.append(", ".join(names) if names else "—")

    row.append(", ".join(nao_vou_names) if nao_vou_names else "—")

    ws.append(row)
    wb.save(HISTORY_FILE)


def history_file_exists():
    return os.path.exists(HISTORY_FILE)


def purge_old_records(max_age_days):
    """
    Remove do histórico as linhas cuja "Data de conclusão" tem mais de
    `max_age_days` dias. Linhas com data em formato inesperado são mantidas
    por segurança. Retorna quantas linhas foram removidas.
    """
    if not history_file_exists():
        return 0

    wb = load_workbook(HISTORY_FILE)
    ws = wb.active
    cutoff = dt.datetime.now() - dt.timedelta(days=max_age_days)

    kept_rows = []
    removed = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or row[0] is None:
            continue
        try:
            row_date = dt.datetime.strptime(str(row[0]), "%d/%m/%Y %H:%M")
        except ValueError:
            kept_rows.append(row)
            continue
        if row_date >= cutoff:
            kept_rows.append(row)
        else:
            removed += 1

    if removed == 0:
        return 0

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "Histórico"
    new_ws.append(FULL_HEADERS)
    for col_idx, header in enumerate(FULL_HEADERS, start=1):
        new_ws.column_dimensions[new_ws.cell(row=1, column=col_idx).column_letter].width = max(
            14, min(40, len(header) + 4)
        )
    for row in kept_rows:
        new_ws.append(row)
    new_wb.save(HISTORY_FILE)

    return removed
